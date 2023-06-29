import sys, openpyxl
from openpyxl.styles import Font, NamedStyle

if len(sys.argv) != 2:
    print('Please enter a number to create a NxN table.')
    sys.exit(0)

tableSize = int(sys.argv[1])

wb = openpyxl.Workbook()
wb.create_sheet(title='MultipliationTable', index=0)
sheet = wb['MultipliationTable']
boldStyle = NamedStyle(name="Bold", font=Font(bold=True))

for row in range(1, tableSize + 2):
    for col in range(1, tableSize + 2):
        if not (row == 1 and col == 1):
            if row == 1:
                sheet.cell(row, col).style = boldStyle
                sheet.cell(row, col).value = col - 1
            elif col == 1:
                sheet.cell(row, col).style = boldStyle
                sheet.cell(row, col).value = row - 1
            else:
                sheet.cell(row, col).value = (row - 1) * (col - 1)

wb.save('MultiplicationTable.xlsx')
